import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ---------------- TYPE MAP ----------------
TYPE_MAP = {
    "G": "Govt",
    "S": "Self",
    "A": "Aided",
    "P": "Private",
}


def get_type_from_code(code: str) -> str:
    """Extract college type from 2nd character of the code."""
    if not code or len(code) < 2:
        return "Unknown"
    return TYPE_MAP.get(code[1].upper(), "Other")


# ---------------- MAIN COMPARISON ----------------
def compare_excels(file1, file2):
    import pandas as pd
    from io import BytesIO
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    # Read both Excel files
    df1 = pd.read_excel(file1, engine="openpyxl")
    df2 = pd.read_excel(file2, engine="openpyxl")

    # Remove unnamed/empty columns
    df1 = df1.loc[:, ~df1.columns.str.contains("^Unnamed")]
    df2 = df2.loc[:, ~df2.columns.str.contains("^Unnamed")]

    # Validate required columns
    required_cols = ["CounselGroup", "CollegeType", "CollegeCode", "CourseCode", "Category", "Seat"]
    for df, name in [(df1, "Input 1"), (df2, "Input 2")]:
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            raise ValueError(f"{name} missing required columns: {', '.join(missing)}")

    # Build unique comparison code
    for df, label in [(df1, "1"), (df2, "2")]:
        df[f"Code{label}"] = (
            df["CounselGroup"].astype(str).str.strip()
            + df["CollegeType"].astype(str).str.strip()
            + df["CollegeCode"].astype(str).str.strip()
            + df["CourseCode"].astype(str).str.strip()
            + df["Category"].astype(str).str.strip()
        )

    # Merge both inputs
    merged = df1.merge(df2, left_on="Code1", right_on="Code2", how="outer", suffixes=("_1", "_2"))

    # Compute difference and status
    merged["Difference"] = merged["Seat_1"].fillna(0) - merged["Seat_2"].fillna(0)

    def get_status(row):
        if pd.isna(row["Seat_2"]):
            return "Only in Input 1"
        elif pd.isna(row["Seat_1"]):
            return "Only in Input 2"
        elif row["Seat_1"] != row["Seat_2"]:
            return "Seat Mismatch"
        else:
            return "Matched"

    merged["Status"] = merged.apply(get_status, axis=1)
    merged["Type"] = merged["Code1"].fillna(merged["Code2"]).apply(get_type_from_code)

    # ---------------- MAIN COMPARISON SHEET ----------------
    comparison_df = merged[
        ["Type", "Code1", "Seat_1", "Code2", "Seat_2", "Difference", "Status"]
    ].rename(columns={"Seat_1": "Input1_Seats", "Seat_2": "Input2_Seats"})

    # ---------------- SEAT DIFFERENCE SHEET ----------------
    seat_diff_df = merged.copy()

    # Fill missing details from Input2 side when not in Input1
    for col in ["CounselGroup", "CollegeType", "CollegeCode", "CourseCode", "Category"]:
        seat_diff_df[f"{col}_1"] = seat_diff_df[f"{col}_1"].combine_first(seat_diff_df[f"{col}_2"])

    seat_diff_df = seat_diff_df[
        [
            "CounselGroup_1",
            "CollegeType_1",
            "CollegeCode_1",
            "CourseCode_1",
            "Category_1",
            "Seat_1",
            "Seat_2",
            "Difference",
            "Status",
        ]
    ].rename(
        columns={
            "CounselGroup_1": "CounselGroup",
            "CollegeType_1": "CollegeType",
            "CollegeCode_1": "CollegeCode",
            "CourseCode_1": "CourseCode",
            "Category_1": "Category",
            "Seat_1": "Input1_Seat",
            "Seat_2": "Input2_Seat",
        }
    )

    # Keep only rows where seat difference is non-zero
    seat_diff_df = seat_diff_df[seat_diff_df["Difference"] <= 0]

    # ---------------- SAVE TO EXCEL ----------------
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        comparison_df.to_excel(writer, index=False, sheet_name="Seat Comparison")
        seat_diff_df.to_excel(writer, index=False, sheet_name="Seat Difference")

    # Highlight mismatched rows in comparison sheet
    output.seek(0)
    wb = load_workbook(output)
    ws = wb["Seat Comparison"]

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    orange_fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        status = ws.cell(row=row, column=7).value
        if status in ["Seat Mismatch", "Only in Input 1", "Only in Input 2"]:
            ws.cell(row=row, column=6).fill = red_fill  # Difference column
            ws.cell(row=row, column=7).fill = orange_fill  # Status column

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return comparison_df, final_output


# ---------------- STREAMLIT UI ----------------
def seat_comparison_ui():
    st.subheader("📊 Excel Seat Comparison Tool")

    st.info("Upload two Excel files with columns: CounselGroup | CollegeType | CollegeCode | CourseCode | Category | Seat")

    col1, col2 = st.columns(2)
    with col1:
        file1 = st.file_uploader("Upload Input Excel 1 - Latest Seat", type=["xlsx", "xls"], key="file1")
    with col2:
        file2 = st.file_uploader("Upload Input Excel 2 - Previous Seat", type=["xlsx", "xls"], key="file2")

    if file1 and file2:
        if st.button("🔍 Run Comparison"):
            with st.spinner("Comparing seats..."):
                try:
                    df_out, excel_out = compare_excels(file1, file2)
                    st.success("✅ Comparison completed!")

                    st.dataframe(df_out, use_container_width=True)

                    st.download_button(
                        "📥 Download Comparison Excel",
                        data=excel_out,
                        file_name="seat_comparison.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                except Exception as e:
                    st.error(f"Error: {e}")
